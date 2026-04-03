from pathlib import Path
from typing import List, Tuple

from openpyxl import load_workbook

from app.engine.masker import MaskingEngine
from app.models.report import HitDetail


def process_xlsx(file_path: Path, output_path: Path, engine: MaskingEngine) -> Tuple[Path, List[HitDetail]]:
    workbook = load_workbook(file_path, data_only=False)
    details: List[HitDetail] = []

    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.data_type == 'f' or (isinstance(cell.value, str) and cell.value.startswith('=')):
                    details.append(HitDetail(
                        entity_type="SKIPPED",
                        rule_type="guard",
                        location=f"{sheet.title}!{cell.coordinate}",
                        original_preview="formula cell",
                        masked_preview="formula cell",
                        confidence=1.0,
                        skipped_reason="formula_cell"
                    ))
                    continue

                if not isinstance(cell.value, str) or not cell.value.strip():
                    continue

                location = f"{sheet.title}!{cell.coordinate}"
                masked_text, hits = engine.desensitize_text(cell.value, location=location)
                if hits:
                    cell.value = masked_text
                    details.extend(hits)

    workbook.save(output_path)
    return output_path, details
